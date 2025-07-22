from flask import render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_login import login_user, logout_user, login_required, current_user
from app import app, db
from models import NetworkAddress, PingLog, NetworkInterface, PingSettings, User, UserRole, AuditLog
from services.network_service import NetworkService
from services.ping_scheduler import scheduler
from auth_decorators import viewer_required, user_required, admin_required, superadmin_required, audit_log, rate_limit
from auth_forms import LoginForm, CreateUserForm, EditUserForm, ChangePasswordForm, ForcePasswordChangeForm, ResetPasswordForm, UnlockUserForm, AuditLogFilterForm
from services.ip_whitelist_service import IPWhitelistService
import logging
import tempfile
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from functools import wraps

logger = logging.getLogger(__name__)

# Декоратор для обработки ошибок маршрутов
def handle_route_errors(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except Exception as e:
            logger.error(f"Error in route {f.__name__}: {str(e)}")
            flash(f'Произошла ошибка: {str(e)}', 'error')
            return redirect(url_for('index'))
    return decorated_function

# ====================== AUTHENTICATION ROUTES ======================

@app.route('/login', methods=['GET', 'POST'])
@rate_limit(max_attempts=5)
def login():
    """Страница входа в систему"""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        
        if user and user.check_password(form.password.data):
            if user.is_locked():
                flash('Ваш аккаунт заблокирован. Обратитесь к администратору.', 'error')
                return render_template('login.html', form=form)
            
            if not user.is_active:
                flash('Ваш аккаунт деактивирован. Обратитесь к администратору.', 'error')
                return render_template('login.html', form=form)
            
            login_user(user, remember=form.remember_me.data)
            user.reset_login_attempts()
            
            # Логирование успешного входа
            AuditLog.log_action(
                action='Вход в систему',
                user_id=user.id,
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent', '')
            )
            
            flash(f'Добро пожаловать, {user.username}!', 'success')
            
            # Check if user needs to change password
            if hasattr(user, 'force_password_change') and user.force_password_change:
                flash('Необходимо изменить пароль по умолчанию', 'warning')
                return redirect(url_for('force_password_change'))
            
            # Redirect to next page or dashboard
            next_page = request.args.get('next')
            if next_page:
                return redirect(next_page)
            return redirect(url_for('index'))
        else:
            # Неудачная попытка входа
            if user:
                user.increment_login_attempts()
            
            AuditLog.log_action(
                action='Неудачная попытка входа',
                details=f'Пользователь: {form.username.data}',
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent', '')
            )
            
            flash('Неверное имя пользователя или пароль', 'error')
    
    return render_template('login.html', form=form)

@app.route('/force_password_change', methods=['GET', 'POST'])
@login_required
def force_password_change():
    """Принудительная смена пароля"""
    form = ForcePasswordChangeForm()
    if form.validate_on_submit():
        if not current_user.check_password(form.current_password.data):
            flash('Неверный текущий пароль', 'error')
            return render_template('force_password_change.html', form=form)
        
        # Validate new password strength
        new_password = form.new_password.data
        if len(new_password) < 8:
            flash('Пароль должен содержать не менее 8 символов', 'error')
            return render_template('force_password_change.html', form=form)
        
        if not any(c.isupper() for c in new_password):
            flash('Пароль должен содержать хотя бы одну заглавную букву', 'error')
            return render_template('force_password_change.html', form=form)
        
        if not any(c.islower() for c in new_password):
            flash('Пароль должен содержать хотя бы одну строчную букву', 'error')
            return render_template('force_password_change.html', form=form)
        
        if not any(c.isdigit() for c in new_password):
            flash('Пароль должен содержать хотя бы одну цифру', 'error')
            return render_template('force_password_change.html', form=form)
        
        # Update password
        current_user.set_password(new_password)
        if hasattr(current_user, 'force_password_change'):
            current_user.force_password_change = False
        db.session.commit()
        
        # Log the action
        AuditLog.log_action(
            action='Принудительная смена пароля',
            user_id=current_user.id,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')
        )
        
        flash('Пароль успешно изменен', 'success')
        return redirect(url_for('index'))
    
    return render_template('force_password_change.html', form=form)

@app.route('/logout')
@login_required
@audit_log('Выход из системы')
def logout():
    """Выход из системы"""
    logout_user()
    flash('Вы вышли из системы', 'info')
    return redirect(url_for('login'))

@app.route('/user_management')
@superadmin_required
def user_management():
    """Управление пользователями"""
    users = User.query.all()
    return render_template('user_management.html', users=users)

@app.route('/create_user', methods=['GET', 'POST'])
@superadmin_required
def create_user():
    """Создание нового пользователя"""
    form = CreateUserForm()
    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            role=UserRole(form.role.data)
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        
        AuditLog.log_action(
            action='Создание пользователя',
            user_id=current_user.id,
            details=f'Создан пользователь: {user.username}, роль: {user.role.value}',
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')
        )
        
        flash(f'Пользователь {user.username} создан успешно', 'success')
        return redirect(url_for('user_management'))
    
    return render_template('create_user.html', form=form)

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@superadmin_required
def edit_user(user_id):
    """Редактирование пользователя"""
    user = User.query.get_or_404(user_id)
    form = EditUserForm(obj=user)
    
    if form.validate_on_submit():
        old_data = f'Роль: {user.role.value}, Активен: {user.is_active}'
        
        user.username = form.username.data
        user.role = UserRole(form.role.data)
        user.is_active = form.is_active.data
        db.session.commit()
        
        AuditLog.log_action(
            action='Изменение пользователя',
            user_id=current_user.id,
            details=f'Изменен пользователь: {user.username}. Было: {old_data}, Стало: Роль: {user.role.value}, Активен: {user.is_active}',
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')
        )
        
        flash(f'Пользователь {user.username} изменен успешно', 'success')
        return redirect(url_for('user_management'))
    
    return render_template('edit_user.html', form=form, user=user)

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Смена пароля текущего пользователя"""
    form = ChangePasswordForm()
    if form.validate_on_submit():
        if current_user.check_password(form.current_password.data):
            current_user.set_password(form.new_password.data)
            db.session.commit()
            
            AuditLog.log_action(
                action='Смена пароля',
                user_id=current_user.id,
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent', '')
            )
            
            flash('Пароль изменен успешно', 'success')
            return redirect(url_for('index'))
        else:
            flash('Неверный текущий пароль', 'error')
    
    return render_template('change_password.html', form=form)

@app.route('/reset_password/<int:user_id>', methods=['GET', 'POST'])
@superadmin_required
def reset_password(user_id):
    """Сброс пароля пользователя администратором"""
    user = User.query.get_or_404(user_id)
    form = ResetPasswordForm()
    
    if form.validate_on_submit():
        user.set_password(form.new_password.data)
        user.unlock_account()  # Разблокировать после сброса пароля
        db.session.commit()
        
        AuditLog.log_action(
            action='Сброс пароля',
            user_id=current_user.id,
            details=f'Сброшен пароль для пользователя: {user.username}',
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')
        )
        
        flash(f'Пароль для пользователя {user.username} сброшен успешно', 'success')
        return redirect(url_for('user_management'))
    
    return render_template('reset_password.html', form=form, user=user)

@app.route('/unlock_user/<int:user_id>', methods=['POST'])
@superadmin_required
def unlock_user(user_id):
    """Разблокировка пользователя"""
    user = User.query.get_or_404(user_id)
    user.unlock_account()
    
    AuditLog.log_action(
        action='Разблокировка пользователя',
        user_id=current_user.id,
        details=f'Разблокирован пользователь: {user.username}',
        ip_address=request.remote_addr,
        user_agent=request.headers.get('User-Agent', '')
    )
    
    flash(f'Пользователь {user.username} разблокирован', 'success')
    return redirect(url_for('user_management'))

@app.route('/delete_user/<int:user_id>', methods=['POST'])
@superadmin_required
def delete_user(user_id):
    """Удаление пользователя"""
    user = User.query.get_or_404(user_id)
    
    # Нельзя удалить самого себя
    if user.id == current_user.id:
        flash('Нельзя удалить свою учетную запись', 'error')
        return redirect(url_for('user_management'))
    
    # Нельзя удалить единственного суперадмина
    if user.role == UserRole.SUPERADMIN:
        superadmin_count = User.query.filter_by(role=UserRole.SUPERADMIN).count()
        if superadmin_count <= 1:
            flash('Нельзя удалить единственного суперадмина', 'error')
            return redirect(url_for('user_management'))
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    
    AuditLog.log_action(
        action='Удаление пользователя',
        user_id=current_user.id,
        details=f'Удален пользователь: {username}',
        ip_address=request.remote_addr,
        user_agent=request.headers.get('User-Agent', '')
    )
    
    flash(f'Пользователь {username} удален', 'success')
    return redirect(url_for('user_management'))

@app.route('/audit_log')
@admin_required
def audit_log():
    """Журнал аудита"""
    form = AuditLogFilterForm()
    
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    query = AuditLog.query.order_by(AuditLog.timestamp.desc())
    
    # Применение фильтров
    username = request.args.get('username')
    action = request.args.get('action')
    
    if username:
        query = query.join(User).filter(User.username.contains(username))
    if action:
        query = query.filter(AuditLog.action.contains(action))
    
    logs = query.paginate(page=page, per_page=per_page, error_out=False)
    
    return render_template('audit_log.html', logs=logs, form=form)

# ====================== MAIN APPLICATION ROUTES ======================

@app.route('/test')
def test():
    """Simple test page"""
    return "<h1>Тестовая страница</h1><p>Если эта страница отображается, Flask работает корректно.</p>"

@app.route('/simple')
def simple():
    """Simple page without external dependencies"""
    addresses = NetworkAddress.query.filter_by(is_active=True).all()
    return render_template('simple.html', addresses=addresses)

@app.route('/')
@viewer_required
@handle_route_errors
def index():
    """Main dashboard showing network addresses and their status"""
    addresses = NetworkAddress.query.filter_by(is_active=True).all()
    
    # Group addresses by group_name
    grouped_addresses = {}
    for address in addresses:
        group_name = address.group_name or 'Основная'
        if group_name not in grouped_addresses:
            grouped_addresses[group_name] = []
        grouped_addresses[group_name].append(address)
    
    return render_template('index.html', addresses=addresses, grouped_addresses=grouped_addresses)

@app.route('/add_address', methods=['POST'])
@user_required
@handle_route_errors
def add_address():
    """Add a new network address to monitor"""
    ip_address = request.form.get('ip_address')
    group_name = request.form.get('group_name', 'Основная')
    
    if not ip_address:
        flash('IP адрес обязателен', 'error')
        return redirect(url_for('index'))
    
    # Check if address already exists
    existing = NetworkAddress.query.filter_by(ip_address=ip_address).first()
    if existing:
        flash('IP адрес уже существует', 'error')
        return redirect(url_for('index'))
    
    try:
        # Validate IP address
        if not NetworkService.validate_ip(ip_address):
            flash('Неверный формат IP адреса', 'error')
            return redirect(url_for('index'))
        
        # Add new address
        new_address = NetworkAddress(
            ip_address=ip_address,
            group_name=group_name
        )
        db.session.add(new_address)
        db.session.commit()
        
        flash('Сетевой адрес успешно добавлен', 'success')
        logger.info(f"Added new network address: {ip_address} to group: {group_name}")
        
    except Exception as e:
        logger.error(f"Error adding network address: {str(e)}")
        flash('Ошибка добавления сетевого адреса', 'error')
        db.session.rollback()
    
    return redirect(url_for('index'))

@app.route('/remove_address/<int:address_id>')
@user_required
def remove_address(address_id):
    """Remove a network address from monitoring"""
    try:
        address = NetworkAddress.query.get_or_404(address_id)
        address.is_active = False
        db.session.commit()
        flash('Network address removed successfully', 'success')
        logger.info(f"Removed network address: {address.ip_address}")
    except Exception as e:
        logger.error(f"Error removing network address: {str(e)}")
        flash('Error removing network address', 'error')
        db.session.rollback()
    
    return redirect(url_for('index'))

@app.route('/logs')
@viewer_required
def logs():
    """Show ping logs with filtering options"""
    address_id = request.args.get('address_id', type=int)
    status_changes_only = request.args.get('status_changes_only', 'false').lower() == 'true'
    limit = request.args.get('limit', 100, type=int)
    
    addresses = NetworkAddress.query.filter_by(is_active=True).all()
    
    if status_changes_only:
        ping_logs = PingLog.get_status_changes(address_id, limit)
    else:
        query = PingLog.query
        if address_id:
            query = query.filter_by(network_address_id=address_id)
        ping_logs = query.order_by(PingLog.timestamp.desc()).limit(limit).all()
    
    return render_template('logs.html', 
                         ping_logs=ping_logs, 
                         addresses=addresses,
                         selected_address_id=address_id,
                         status_changes_only=status_changes_only,
                         limit=limit)

@app.route('/settings')
@admin_required
def settings():
    """Network interface settings and monitoring configuration"""
    # Get current ping settings
    ping_settings = PingSettings.get_current()
    
    return render_template('settings.html', ping_settings=ping_settings)



@app.route('/ping_now/<int:address_id>')
@user_required
def ping_now(address_id):
    """Manually trigger a ping for a specific address"""
    try:
        address = NetworkAddress.query.get_or_404(address_id)
        network_service = NetworkService()
        
        # Perform ping
        result = network_service.ping_address(address.ip_address)
        
        # Update address status
        address.last_status = result['status']
        address.last_ping_time = result['timestamp']
        
        # Log the result
        ping_log = PingLog(
            network_address_id=address.id,
            status=result['status'],
            response_time=result['response_time'],
            error_message=result.get('error_message')
        )
        
        db.session.add(ping_log)
        db.session.commit()
        
        flash(f'Ping completed: {address.ip_address} is {result["status"]}', 'success')
        logger.info(f"Manual ping completed for {address.ip_address}: {result['status']}")
        
    except Exception as e:
        logger.error(f"Error performing manual ping: {str(e)}")
        flash('Error performing ping', 'error')
        db.session.rollback()
    
    return redirect(url_for('index'))

@app.route('/api/status')
@viewer_required
def api_status():
    """API endpoint for getting current status of all addresses"""
    addresses = NetworkAddress.query.filter_by(is_active=True).all()
    return jsonify([addr.to_dict() for addr in addresses])

@app.route('/scheduler/start')
@admin_required
def start_scheduler():
    """Start the ping scheduler"""
    try:
        from services.ping_scheduler import start_scheduler
        start_scheduler()
        flash('Scheduler started', 'success')
    except Exception as e:
        logger.error(f"Error starting scheduler: {str(e)}")
        flash('Error starting scheduler', 'error')
    
    return redirect(url_for('settings'))

@app.route('/scheduler/stop')
@admin_required
def stop_scheduler():
    """Stop the ping scheduler"""
    try:
        if scheduler and scheduler.running:
            scheduler.shutdown()
            flash('Планировщик остановлен', 'success')
        else:
            flash('Планировщик не запущен', 'info')
    except Exception as e:
        logger.error(f"Error stopping scheduler: {str(e)}")
        flash('Ошибка остановки планировщика', 'error')
    
    return redirect(url_for('settings'))

@app.route('/update_ping_settings', methods=['POST'])
@admin_required
def update_ping_settings():
    """Update ping settings including async parameters"""
    try:
        ping_interval = request.form.get('ping_interval', 30, type=int)
        timeout = request.form.get('timeout', 5, type=int)
        max_retries = request.form.get('max_retries', 3, type=int)
        max_threads = request.form.get('max_threads', 50, type=int)
        batch_size = request.form.get('batch_size', 100, type=int)
        
        # Validation
        if ping_interval < 5 or ping_interval > 3600:
            flash('Интервал пинга должен быть от 5 до 3600 секунд', 'error')
            return redirect(url_for('settings'))
            
        if timeout < 1 or timeout > 30:
            flash('Таймаут должен быть от 1 до 30 секунд', 'error')
            return redirect(url_for('settings'))
            
        if max_retries < 1 or max_retries > 10:
            flash('Количество попыток должно быть от 1 до 10', 'error')
            return redirect(url_for('settings'))
            
        if max_threads < 1 or max_threads > 200:
            flash('Количество потоков должно быть от 1 до 200', 'error')
            return redirect(url_for('settings'))
            
        if batch_size < 10 or batch_size > 1000:
            flash('Размер пакета должен быть от 10 до 1000', 'error')
            return redirect(url_for('settings'))
        
        # Update settings
        settings = PingSettings.get_current()
        settings.ping_interval = ping_interval
        settings.timeout = timeout
        settings.max_retries = max_retries
        settings.max_threads = max_threads
        settings.batch_size = batch_size
        
        db.session.commit()
        
        # Restart scheduler with new settings
        from services.ping_scheduler import restart_scheduler
        restart_scheduler()
        
        flash('Настройки пинга обновлены', 'success')
        logger.info(f"Updated ping settings: interval={ping_interval}s, timeout={timeout}s, retries={max_retries}, threads={max_threads}, batch={batch_size}")
        
    except Exception as e:
        logger.error(f"Error updating ping settings: {str(e)}")
        flash('Ошибка обновления настроек', 'error')
        db.session.rollback()
    
    return redirect(url_for('settings'))

@app.route('/optimize_ping_settings', methods=['POST'])
@admin_required
def optimize_ping_settings():
    """Auto-optimize ping settings based on current address count"""
    try:
        from services.async_ping_service import AsyncPingService
        
        # Get current address count
        address_count = NetworkAddress.query.filter_by(is_active=True).count()
        
        # Get recommended settings
        async_service = AsyncPingService()
        recommended = async_service.get_recommended_settings(address_count)
        
        # Update settings
        settings = PingSettings.get_current()
        settings.max_threads = recommended['max_threads']
        settings.batch_size = recommended['batch_size']
        
        db.session.commit()
        
        flash(f'Настройки оптимизированы для {address_count} адресов: {recommended["max_threads"]} потоков, пакет {recommended["batch_size"]}', 'success')
        logger.info(f"Auto-optimized settings for {address_count} addresses: {recommended}")
        
    except Exception as e:
        logger.error(f"Error optimizing ping settings: {str(e)}")
        flash('Ошибка автооптимизации', 'error')
        db.session.rollback()
    
    return redirect(url_for('settings'))

@app.route('/export_database')
@admin_required
def export_database():
    """Export database to Excel file"""
    try:
        # Create workbook
        wb = Workbook()
        
        # Export Network Addresses
        ws1 = wb.active
        ws1.title = "Network Addresses"
        ws1.append(['ID', 'IP Address', 'Group Name', 'Active', 'Last Status', 'Last Ping Time', 'Created At'])
        
        addresses = NetworkAddress.query.all()
        for addr in addresses:
            ws1.append([
                addr.id,
                addr.ip_address,
                addr.group_name,
                addr.is_active,
                addr.last_status,
                addr.last_ping_time.isoformat() if addr.last_ping_time else None,
                addr.created_at.isoformat()
            ])
        
        # Export Ping Logs
        ws2 = wb.create_sheet("Ping Logs")
        ws2.append(['ID', 'Network Address ID', 'IP Address', 'Status', 'Response Time', 'Timestamp', 'Error Message'])
        
        logs = PingLog.query.join(NetworkAddress).order_by(PingLog.timestamp.desc()).limit(10000).all()
        for log in logs:
            ws2.append([
                log.id,
                log.network_address_id,
                log.network_address.ip_address,
                log.status,
                log.response_time,
                log.timestamp.isoformat(),
                log.error_message
            ])
        
        # Export Network Interfaces
        ws3 = wb.create_sheet("Network Interfaces")
        ws3.append(['ID', 'Name', 'IP Address', 'Subnet', 'Selected', 'Created At'])
        
        interfaces = NetworkInterface.query.all()
        for iface in interfaces:
            ws3.append([
                iface.id,
                iface.name,
                iface.ip_address,
                iface.subnet,
                iface.is_selected,
                iface.created_at.isoformat()
            ])
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'network_monitor_export_{timestamp}.xlsx'
        
        return send_file(temp_file.name, as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"Error exporting database: {str(e)}")
        flash('Ошибка экспорта базы данных', 'error')
        return redirect(url_for('settings'))

@app.route('/import_database', methods=['POST'])
@admin_required
def import_database():
    """Import network addresses from Excel file"""
    try:
        if 'file' not in request.files:
            flash('Файл не выбран', 'error')
            return redirect(url_for('settings'))
        
        file = request.files['file']
        if file.filename == '':
            flash('Файл не выбран', 'error')
            return redirect(url_for('settings'))
        
        if not file.filename.endswith('.xlsx'):
            flash('Поддерживаются только файлы .xlsx', 'error')
            return redirect(url_for('settings'))
        
        # Save uploaded file temporarily
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        
        # Load workbook
        wb = load_workbook(temp_file.name)
        
        # Import from Network Addresses sheet
        if 'Network Addresses' in wb.sheetnames:
            ws = wb['Network Addresses']
            imported_count = 0
            
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1]:  # IP Address column
                    ip_address = str(row[1])
                    group_name = str(row[2]) if row[2] else "Основная"
                    
                    # Check if address already exists
                    existing = NetworkAddress.query.filter_by(ip_address=ip_address).first()
                    if not existing:
                        # Validate IP address
                        if NetworkService.validate_ip(ip_address):
                            new_address = NetworkAddress(
                                ip_address=ip_address,
                                group_name=group_name
                            )
                            db.session.add(new_address)
                            imported_count += 1
            
            db.session.commit()
            flash(f'Импортировано {imported_count} сетевых адресов', 'success')
            logger.info(f"Imported {imported_count} network addresses")
        else:
            flash('Не найден лист "Network Addresses"', 'error')
        
        # Clean up temp file
        os.unlink(temp_file.name)
        
    except Exception as e:
        logger.error(f"Error importing database: {str(e)}")
        flash('Ошибка импорта базы данных', 'error')
        db.session.rollback()
    
    return redirect(url_for('settings'))

# ====================== IP WHITELIST MANAGEMENT ======================

@app.route('/ip_whitelist')
@admin_required
def ip_whitelist():
    """Управление белым списком IP-адресов"""
    whitelist_service = IPWhitelistService()
    
    config = whitelist_service.get_config_info()
    whitelist = whitelist_service.get_whitelist()
    groups = whitelist_service.get_groups()
    client_ip = whitelist_service.get_client_ip()
    
    return render_template('ip_whitelist.html', 
                         config=config, 
                         whitelist=whitelist,
                         groups=groups,
                         client_ip=client_ip,
                         user_role=current_user.role.value)

@app.route('/ip_whitelist/add', methods=['POST'])
@admin_required
def add_ip_whitelist():
    """Добавление IP-адреса в белый список"""
    whitelist_service = IPWhitelistService()
    
    ip_address = request.form.get('ip_address', '').strip()
    
    if not ip_address:
        return jsonify({'success': False, 'message': 'IP-адрес не указан'})
    
    if whitelist_service.add_ip(ip_address, current_user.username):
        AuditLog.log_action(
            action='Добавлен IP в белый список',
            details=f'IP-адрес: {ip_address}',
            user_id=current_user.id,
            ip_address=request.remote_addr
        )
        
        # Send real-time update via WebSocket (только для того, кто добавляет)
        try:
            from app import socketio
            config = whitelist_service.get_config_info()
            whitelist = whitelist_service.get_whitelist()
            groups = whitelist_service.get_groups()
            
            # Отправляем уведомление только текущему пользователю
            socketio.emit('whitelist_update', {
                'action': 'add',
                'ip_address': ip_address,
                'config': config,
                'whitelist': whitelist,
                'groups': groups,
                'message': 'IP-адрес добавлен в белый список'
            }, room=f'user_{current_user.id}')
        except Exception as e:
            logger.error(f"Error sending whitelist WebSocket update: {str(e)}")
        
        return jsonify({'success': True, 'message': 'IP-адрес добавлен в белый список'})
    else:
        return jsonify({'success': False, 'message': 'Неверный формат IP-адреса или адрес уже существует'})

@app.route('/ip_whitelist/remove', methods=['POST'])
@admin_required
def remove_ip_whitelist():
    """Удаление IP-адреса из белого списка"""
    whitelist_service = IPWhitelistService()
    
    ip_address = request.form.get('ip_address', '').strip()
    
    if not ip_address:
        return jsonify({'success': False, 'message': 'IP-адрес не указан'})
    
    if whitelist_service.remove_ip(ip_address, current_user.username, current_user.role.value):
        AuditLog.log_action(
            action='Удален IP из белого списка',
            details=f'IP-адрес: {ip_address}',
            user_id=current_user.id,
            ip_address=request.remote_addr
        )
        
        # Send real-time update via WebSocket (только для того, кто удаляет)
        try:
            from app import socketio
            config = whitelist_service.get_config_info()
            whitelist = whitelist_service.get_whitelist()
            groups = whitelist_service.get_groups()
            
            # Отправляем уведомление только текущему пользователю
            socketio.emit('whitelist_update', {
                'action': 'remove',
                'ip_address': ip_address,
                'config': config,
                'whitelist': whitelist,
                'groups': groups,
                'message': 'IP-адрес удален из белого списка'
            }, room=f'user_{current_user.id}')
        except Exception as e:
            logger.error(f"Error sending whitelist WebSocket update: {str(e)}")
        
        return jsonify({'success': True, 'message': 'IP-адрес удален из белого списка'})
    else:
        return jsonify({'success': False, 'message': 'IP-адрес не найден в белом списке или у вас нет прав на его удаление'})

@app.route('/ip_whitelist/toggle', methods=['POST'])
@superadmin_required
def toggle_ip_whitelist():
    """Включение/выключение белого списка"""
    whitelist_service = IPWhitelistService()
    
    enabled = request.form.get('enabled', 'true').lower() == 'true'
    
    whitelist_service.toggle_whitelist(enabled, current_user.username)
    
    AuditLog.log_action(
        action='Изменен статус белого списка IP',
        details=f'Белый список: {"включен" if enabled else "отключен"}',
        user_id=current_user.id,
        ip_address=request.remote_addr
    )
    
    # Send real-time update via WebSocket (только для того, кто переключает)
    try:
        from app import socketio
        config = whitelist_service.get_config_info()
        whitelist = whitelist_service.get_whitelist()
        groups = whitelist_service.get_groups()
        
        # Отправляем уведомление только текущему пользователю
        socketio.emit('whitelist_update', {
            'action': 'toggle',
            'enabled': enabled,
            'config': config,
            'whitelist': whitelist,
            'groups': groups,
            'message': f'Белый список {"включен" if enabled else "отключен"}'
        }, room=f'user_{current_user.id}')
    except Exception as e:
        logger.error(f"Error sending whitelist WebSocket update: {str(e)}")
    
    return jsonify({'success': True, 'message': f'Белый список {"включен" if enabled else "отключен"}'})

@app.route('/ip_whitelist/refresh', methods=['POST'])
@superadmin_required
def refresh_ip_whitelist():
    """Обновление белого списка из файла"""
    try:
        whitelist_service = IPWhitelistService()
        
        # Force reload configuration from file
        config = whitelist_service.load_config()
        whitelist = whitelist_service.get_whitelist()
        groups = whitelist_service.get_groups()
        
        # Send real-time update via WebSocket (только для текущего пользователя)
        from app import socketio
        socketio.emit('whitelist_update', {
            'action': 'refresh',
            'config': whitelist_service.get_config_info(),
            'whitelist': whitelist,
            'groups': groups,
            'message': 'Белый список обновлен из файла'
        }, room=f'user_{current_user.id}')
        
        return jsonify({'success': True, 'message': 'Белый список успешно обновлен'})
    except Exception as e:
        logger.error(f"Error refreshing whitelist: {str(e)}")
        return jsonify({'success': False, 'message': f'Ошибка обновления: {str(e)}'})
